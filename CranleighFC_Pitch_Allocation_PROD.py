#!/usr/bin/env python
# coding: utf-8

# In[1]:


"""
Enhanced Cranleigh FC Pitch Allocation System V52.1
FIXED: Properly handles multiple fixtures per team across different dates
Each fixture is uniquely identified by team+date combination
"""

import pandas as pd
from ortools.sat.python import cp_model
from typing import Dict, List, Tuple
import webbrowser
import os
import sys

# ✅ Print environment info at startup
print(f"Python: {sys.executable}")
print(f"Version: {sys.version}")
print(f"Checking openpyxl availability...")
try:
    import openpyxl
    print(f"✓ openpyxl {openpyxl.__version__} is available\n")
except ImportError:
    print(f"✗ openpyxl NOT available in this environment\n")
    print(f"Install with: {sys.executable} -m pip install openpyxl\n")

# =====================================
# 🏟️ Define Pitches
# =====================================
pitches = {
    'P1 11v11 (Bruce McKenzie)': {'format': '11v11', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P2 11v11': {'format': '11v11', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P3 11v11 (Middle)': {'format': '11v11', 'lights': False, 'size': 'small', 'location': 'snoxhall', 'priority': 1},
    'P4 9v9': {'format': '9v9', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P5 7v7': {'format': '7v7', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P6 11v11 (Seniors)': {'format': '11v11', 'lights': True, 'location': 'snoxhall', 'priority': 1},
    'P7 9v9': {'format': '9v9', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P8 7v7': {'format': '7v7', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'P9 7v7': {'format': '7v7', 'lights': False, 'location': 'snoxhall', 'priority': 1},
    'CCC1 5v5': {'format': '5v5', 'lights': False, 'location': 'ccc', 'priority': 1},
    'CCC2 5v5': {'format': '5v5', 'lights': False, 'location': 'ccc', 'priority': 1},
    'CCC3 7v7': {'format': '7v7', 'lights': False, 'location': 'ccc', 'priority': 1},
    'CCC4 7v7': {'format': '7v7', 'lights': False, 'location': 'ccc', 'priority': 1},
    # ✅ NEW: Glebelands 3G pitches (secondary - use after main pitches)
    'G1 11v11 (Glebelands 3G)': {'format': '11v11', 'lights': True, 'location': 'glebelands', 'priority': 2},
    'G2 9v9 (Glebelands 3G)': {'format': '9v9', 'lights': True, 'location': 'glebelands', 'priority': 2}
}

# =====================================
# 🧑 Team and Age Group Setup
# =====================================
valid_teams = {    
    'Cranleigh': 'Seniors',
    'Cranleigh Blues': 'U17',
    'Cranleigh Development': 'Seniors',
    'Cranleigh Dons U14': 'U14',
    'Cranleigh (First)': 'Seniors',
    'Cranleigh Harriers U13': 'U13',
    'Cranleigh Hawks U12': 'U12',
    'Cranleigh Kangaroos U7': 'U7',
    'Cranleigh Koalas U7': 'U7',
    'Cranleigh Kookaburras U7': 'U7',
    'Cranleigh Masters': 'Seniors',
    'Cranleigh Reserves': 'Seniors',
    'Cranleigh U10 Cobras': 'U10',
    'Cranleigh U10 Cyclones': 'U10',
    'Cranleigh U10 Tigers': 'U10',
    'Cranleigh U10 Vipers': 'U10',
    'Cranleigh U10 Wolves': 'U10',
    'Cranleigh U11': 'U11',
    'Cranleigh U11 Crushers': 'U11',
    'Cranleigh U11 Jaguars': 'U11',
    'Cranleigh U11 Leopards': 'U11',
    'Cranleigh U11 Panthers': 'U11',
    'Cranleigh U11 Tigers Girls': 'U11',
    'Cranleigh U12M Harriers': 'U12',
    'Cranleigh U13 Cobras': 'U13',
    'Cranleigh U13 Cosmos Blue': 'U13G',  # ✅ Changed to U13G for girls format
    'Cranleigh U13 Cosmos White': 'U13G',  # ✅ Changed to U13G for girls format
    'Cranleigh U13 Jaguars': 'U13',
    'Cranleigh U14 Atletico': 'U14',
    'Cranleigh U14 Girls': 'U14',
    'Cranleigh U14M Albion': 'U14',
    'Cranleigh U15 Cobras': 'U15',
    'Cranleigh U15 Cranes': 'U15',
    'Cranleigh U16 Sharks': 'U16',
    'Cranleigh U16M Tigers': 'U16',
    'Cranleigh U17 County': 'U17',
    'Cranleigh U8 Barracudas': 'U8',
    'Cranleigh U8 Carnage': 'U8',
    'Cranleigh U8 Rays': 'U8',
    'Cranleigh U8 Sharks': 'U8',
    'Cranleigh U9 Bears': 'U9',
    'Cranleigh U9 Coyotes': 'U9',
    'Cranleigh U9 Cuckoos': 'U9',
    'Cranleigh U9 Eagles': 'U9',
    'Cranleigh U9 Lions': 'U9',
    'Cranleigh U9 Raptors': 'U9',
    'Cranleigh Veterans': 'Seniors',
    'Cranleigh Womens': 'Womens'
}

age_group_formats = {
    'U7': '5v5', 'U8': '5v5',
    'U9': '7v7', 'U10': '7v7',
    'U11': '9v9', 'U12': '9v9',
    'U13': '11v11', 
    'U13G': '9v9',  # ✅ NEW: U13 Girls play 9v9 format
    'U14': '11v11',
    'U15': '11v11', 'U16': '11v11',
    'U17': '11v11', 'U18': '11v11',
    'Seniors': '11v11',
    'Womens': '11v11'
}

age_priority = {
    'U7': 1, 'U8': 2, 'U9': 3, 'U10': 4, 'U11': 5, 'U12': 6,
    'U13': 7, 
    'U13G': 7,  # ✅ NEW: Same priority as U13 boys
    'U14': 8, 'U15': 9, 'U16': 10, 'U17': 11, 'U18': 12,
    'Seniors': 13, 'Womens': 13
}

# ✅ P3 Middle pitch priority - younger 11v11 teams preferred
# Higher number = higher priority for P3
p3_middle_priority = {
    'U13': 3,  # Highest priority for smaller pitch
    'U14': 2,  # Second priority
    'U15': 1,  # Lower priority
    'U16': 0,  # No priority (can use if needed)
    'U17': 0,
    'Seniors': 0,
    'Womens': 0
}

# ✅ Senior team priority for P6 allocation
senior_team_priority = {
    'Cranleigh (First)': 4,
    'Cranleigh Reserves': 3,
    'Cranleigh Development': 2,
    'Cranleigh Veterans': 1,
    'Cranleigh Masters': 1,
    'Cranleigh': 1,
    'Cranleigh Womens': 3
}

# =====================================
# 📅 Load Fixtures and Validate
# =====================================
def load_and_validate_fixtures(filepath: str) -> Tuple[Dict, Dict]:
    """Load fixtures with validation - returns fixtures dict and slots by date"""
    fixtures_df = pd.read_csv(filepath)

    # Rename new columns to the expected names used by the allocator
    fixtures_df = fixtures_df.rename(columns={
        'match_date': 'date',
        'match_time': 'time',
        'home_team_clean': 'team_name'
    })

    # ✅ FIX: Convert time objects to string format 'HH:MM'
    def format_time(time_val):
        """Convert time to HH:MM string format"""
        if pd.isna(time_val):
            return None
        if isinstance(time_val, str):
            parts = time_val.split(':')
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        if hasattr(time_val, 'hour') and hasattr(time_val, 'minute'):
            return f"{time_val.hour:02d}:{time_val.minute:02d}"
        return None
    
    fixtures_df['time'] = fixtures_df['time'].apply(format_time)
    fixtures_df["preferred_time"] = fixtures_df["time"]
    
    # ✅ NEW: Check for Cup fixtures in prefix column
    def is_cup_fixture(prefix):
        """Check if fixture is a cup match"""
        if pd.isna(prefix):
            return False
        return 'cup' in str(prefix).lower()
    
    fixtures_df['is_cup'] = fixtures_df['prefix'].apply(is_cup_fixture)
    
    result=fixtures_df.head(10)
    print(result)
    print(f"\n📋 CSV contains {len(fixtures_df)} fixture records")
    
    # ✅ NEW: Report Cup fixtures
    cup_count = fixtures_df['is_cup'].sum()
    if cup_count > 0:
        print(f"🏆 Found {cup_count} Cup fixtures (priority for 09:30 kickoff)")
    
    # Validation
    errors = []
    for idx, row in fixtures_df.iterrows():
        team = row['team_name']
        if team not in valid_teams:
            errors.append(f"Row {idx}: Unknown team '{team}'")
        if pd.isna(row['date']) or pd.isna(row['time']):
            errors.append(f"Row {idx}: Missing date/time for {team}")
    
    if errors:
        print("⚠️ Validation Errors Found:")
        for err in errors:
            print(f"  - {err}")
        raise ValueError("Fix validation errors before proceeding")
    
    # ✅ Build fixtures dict using UNIQUE KEY per fixture (team + date)
    fixtures = {}
    
    for idx, row in fixtures_df.iterrows():
        team = row['team_name']
        age = valid_teams[team]
        fixture_date = row['date']
        is_cup = row['is_cup']
        
        # ✅ Create unique fixture ID: team_date
        fixture_id = f"{team}_{fixture_date}"
        
        # Get senior team priority (0 for non-seniors)
        team_priority = senior_team_priority.get(team, 0)
        
        # ✅ NEW: For Cup fixtures, preferred time is 09:30
        preferred_time = '09:30' if is_cup else row['time']
        
        fixtures[fixture_id] = {
            'team_name': team,
            'fixture_date': fixture_date,
            'preferred_time': preferred_time,  # ✅ Cup fixtures prefer 09:30
            'original_time': row['time'],       # ✅ Keep original for reference
            'format_req': age_group_formats[age],
            'age_group': age,
            'priority': age_priority[age],
            'pref_pitch': 'P6 11v11 (Seniors)' if age in ['Seniors', 'Womens'] else None,
            'senior_priority': team_priority,
            'is_cup': is_cup  # ✅ NEW: Flag cup fixtures
        }
    
    print(f"✅ Processing {len(fixtures)} unique fixtures")
    
    # ... rest of function stays the same
    
    # Generate available slots BY DATE
    slot_dates = sorted(fixtures_df['date'].unique())
    slots_by_date = {}
    
    for date in slot_dates:
        date_slots = []
        for pitch, info in pitches.items():
            fmt = info['format']
            if fmt == '11v11':
                times = ['09:30', '11:00', '14:00']
            elif fmt in ['9v9', '7v7', '5v5']:
                times = ['09:30', '11:00']
            for t in times:
                date_slots.append((date, t, pitch))
        slots_by_date[date] = date_slots
    
    total_slots = sum(len(slots) for slots in slots_by_date.values())
    print(f'✅ Generated {total_slots} slots across {len(slot_dates)} dates')
    
    # Show fixtures per date breakdown
    print(f'\n📅 Fixtures per date:')
    date_counts = fixtures_df.groupby('date').size().sort_index()
    for date, count in date_counts.items():
        print(f"  {date}: {count} fixtures")
    
    return fixtures, slots_by_date

# =====================================
# ⚙️ Build and Solve Model
# =====================================
def solve_allocation(fixtures: Dict, slots_by_date: Dict, timeout: int = 30):
    """Build and solve the CP-SAT model"""
    model = cp_model.CpModel()
    fixture_slot_vars = {}
    
    # Track reasons why fixtures can't be allocated
    no_slots_teams = []
    constraint_blocked = {}
    
    # Create variables - ONLY for slots on the fixture's scheduled date
    for fixture_id, fdata in fixtures.items():
        fixture_date = fdata['fixture_date']
        
        if fixture_date not in slots_by_date:
            no_slots_teams.append((fixture_id, fixture_date, "Date not in schedule"))
            continue
            
        available_slots = slots_by_date[fixture_date]
        fixture_valid_slots = 0
        blocked_reasons = []
        
        for (date, time, pitch) in available_slots:
            # Track why slots are rejected
            if fdata['format_req'] != pitches[pitch]['format']:
                blocked_reasons.append(f"Format mismatch (needs {fdata['format_req']})")
                continue
            if fdata['age_group'] in ['Seniors', 'Womens'] and time != '14:00':
                blocked_reasons.append(f"Seniors must play at 14:00")
                continue
            if fdata['age_group'] not in ['Seniors', 'Womens'] and time == '14:00':
                blocked_reasons.append(f"Youth can't play at 14:00")
                continue
            
            # Hard constraint - only seniors/womens can use P6
            if pitch == 'P6 11v11 (Seniors)' and fdata['age_group'] not in ['Seniors', 'Womens']:
                blocked_reasons.append(f"Non-seniors can't use P6")
                continue
            
            # This slot is valid!
            var = model.NewBoolVar(f'{fixture_id}_{date}_{time}_{pitch}')
            fixture_slot_vars[(fixture_id, date, time, pitch)] = var
            fixture_valid_slots += 1
        
        # Track fixtures with no valid slots
        if fixture_valid_slots == 0:
            unique_reasons = list(set(blocked_reasons))[:3]
            constraint_blocked[fixture_id] = (fixture_date, unique_reasons)
    
    print(f'\nCreated {len(fixture_slot_vars)} decision variables')
    
    # Report fixtures that can't be allocated
    if no_slots_teams:
        print(f"\n❌ {len(no_slots_teams)} fixtures have no slots on their date:")
        for fixture_id, date, reason in no_slots_teams[:5]:
            print(f"  - {fixture_id}: {reason}")
    
    if constraint_blocked:
        print(f"\n❌ {len(constraint_blocked)} fixtures blocked by constraints:")
        for fixture_id, (date, reasons) in list(constraint_blocked.items())[:5]:
            print(f"  - {fixture_id}")
            for reason in reasons[:2]:
                print(f"    → {reason}")
    
    # ✅ SOFT Constraint: Each fixture assigned AT MOST once (not exactly once)
    # This allows the solver to find a solution even if some fixtures can't be allocated
    allocation_vars = {}  # Track if each fixture is allocated
    
    for fixture_id in fixtures.keys():
        fixture_vars = [v for (fid, _, _, _), v in fixture_slot_vars.items() if fid == fixture_id]
        if fixture_vars:
            # Create indicator variable: is this fixture allocated?
            allocated = model.NewBoolVar(f'allocated_{fixture_id}')
            model.Add(sum(fixture_vars) == 1).OnlyEnforceIf(allocated)
            model.Add(sum(fixture_vars) == 0).OnlyEnforceIf(allocated.Not())
            allocation_vars[fixture_id] = allocated
    
    # Track fixtures with no valid slots at all
    impossible_fixtures = set(fixtures.keys()) - set(allocation_vars.keys())
    
    # Constraint: One fixture per exact time slot (date+time+pitch)
    unique_slots = set((d, t, p) for (_, d, t, p) in fixture_slot_vars.keys())
    for date, time, pitch in unique_slots:
        vars_slot = [v for (_, d, t, p), v in fixture_slot_vars.items() 
                    if d == date and t == time and p == pitch]
        if vars_slot:
            model.Add(sum(vars_slot) <= 1)
    
    # Constraint: Max 2 games per pitch per day
    for date in set(d for (_, d, _, _) in fixture_slot_vars.keys()):
        for pitch in pitches.keys():
            vars_day = [v for (fid, d, _, p), v in fixture_slot_vars.items() 
                       if d == date and p == pitch]
            if vars_day:
                model.Add(sum(vars_day) <= 2)
    
    # ✅ Avoid back-to-back matches on same pitch (any day)
    # Penalize consecutive 09:30 + 11:00 slots on same pitch
    # Glebelands pitches exempt as they're designed for overflow
    backtoback_penalty_vars = []
    for date in set(d for (_, d, _, _) in fixture_slot_vars.keys()):
        for pitch, pitch_info in pitches.items():
            # Skip Glebelands pitches - they can handle back-to-back as overflow
            if pitch_info.get('location') == 'glebelands':
                continue
                
            # Check for back-to-back slots (09:30 + 11:00) on main pitches
            slot_0930 = [(fid, var) for (fid, d, t, p), var in fixture_slot_vars.items() 
                        if d == date and p == pitch and t == '09:30']
            slot_1100 = [(fid, var) for (fid, d, t, p), var in fixture_slot_vars.items() 
                        if d == date and p == pitch and t == '11:00']
            
            if slot_0930 and slot_1100:
                # Create penalty variable: 1 if both slots used on same pitch
                penalty = model.NewBoolVar(f'backtoback_penalty_{date}_{pitch}')
                
                # Penalty is 1 only if both 09:30 AND 11:00 slots are used
                vars_0930 = [v for _, v in slot_0930]
                vars_1100 = [v for _, v in slot_1100]
                
                # both_used = 1 if (09:30 slot used) AND (11:00 slot used)
                model.Add(sum(vars_0930) + sum(vars_1100) == 2).OnlyEnforceIf(penalty)
                model.Add(sum(vars_0930) + sum(vars_1100) <= 1).OnlyEnforceIf(penalty.Not())
                
                backtoback_penalty_vars.append(penalty)
    
    # Objective: Maximize number of allocated fixtures + weighted satisfaction
    objective_terms = []
    
    # Primary goal: maximize number of fixtures allocated
    for fixture_id, allocated_var in allocation_vars.items():
        objective_terms.append(allocated_var * 10000)  # Very high weight for allocation
    
    # ✅ Penalize back-to-back matches on same pitch (any day)
    for penalty_var in backtoback_penalty_vars:
        objective_terms.append(penalty_var * -500)  # Subtract 500 points for back-to-back
    
    # Secondary goal: optimize quality of allocations
    for (fixture_id, date, time, pitch), var in fixture_slot_vars.items():
        f = fixtures[fixture_id]
        weight = f['priority'] * 10  # Reduced weight (secondary to allocation)
        
        # ✅ Penalize Glebelands pitches to make them secondary choice
        pitch_info = pitches[pitch]
        if pitch_info.get('location') == 'glebelands':
            weight -= 300  # Strong penalty - only use as overflow
        
        # ✅ NEW: Cup fixtures get strong bonus for 09:30 kickoff
        if f.get('is_cup', False) and time == '09:30':
            weight += 500  # Very strong preference for 09:30
        
        # ✅ NEW: Cup fixtures get pitch priority (prefer best pitches)
        if f.get('is_cup', False):
            # Prefer main 11v11 pitches for cup games
            if f['format_req'] == '11v11' and pitch in ['P1 11v11 (Bruce McKenzie)', 'P2 11v11', 'P6 11v11 (Seniors)']:
                weight += 200  # Strong preference for premier pitches
            # Prefer main format-appropriate pitches for other cup games
            elif pitch in ['P4 9v9', 'P7 9v9'] and f['format_req'] == '9v9':
                weight += 150
            elif pitch in ['P5 7v7', 'P8 7v7', 'P9 7v7', 'CCC3 7v7', 'CCC4 7v7'] and f['format_req'] == '7v7':
                weight += 150
            elif pitch in ['CCC1 5v5', 'CCC2 5v5'] and f['format_req'] == '5v5':
                weight += 150
        
        # Strong bonus for U13/U14 getting P3 Middle pitch
        if pitch == 'P3 11v11 (Middle)' and f['age_group'] in p3_middle_priority:
            priority_bonus = p3_middle_priority[f['age_group']]
            weight += priority_bonus * 75  # U13 gets +225, U14 gets +150, U15 gets +75
        
        # Strong bonus for senior teams getting P6 based on their priority
        if pitch == 'P6 11v11 (Seniors)' and f['senior_priority'] > 0:
            weight += f['senior_priority'] * 50
        
        # Bonus for preferred time (including Cup 09:30 preference)
        if time == f['preferred_time']:
            weight += 50
        
        objective_terms.append(weight * var)
    
    model.Maximize(sum(objective_terms))
    
    # Solve
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = timeout
    solver.parameters.num_search_workers = 8
    
    print(f'\n🔍 Solving (timeout: {timeout}s)...')
    status = solver.Solve(model)
    
    # Report results
    status_map = {
        cp_model.OPTIMAL: '✅ OPTIMAL',
        cp_model.FEASIBLE: '✓ FEASIBLE (not proven optimal)',
        cp_model.INFEASIBLE: '❌ INFEASIBLE',
        cp_model.MODEL_INVALID: '❌ INVALID MODEL',
        cp_model.UNKNOWN: '⚠️ UNKNOWN'
    }
    
    print(f'\nStatus: {status_map.get(status, "UNKNOWN")}')
    print(f'Wall time: {solver.WallTime():.2f}s')
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print(f'Objective value: {solver.ObjectiveValue()}')
    
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # Extract solution
        allocations = []
        allocated_fixture_ids = set()
        
        for (fixture_id, date, time, pitch), var in fixture_slot_vars.items():
            if solver.Value(var) == 1:
                allocated_fixture_ids.add(fixture_id)
                f = fixtures[fixture_id]
                matched_time = time == f['preferred_time']
                matched_pitch = pitch == f['pref_pitch'] if f['pref_pitch'] else False
                
                allocations.append({
                    'fixture_id': fixture_id,
                    'team': f['team_name'],
                    'date': date,
                    'time': time,
                    'pitch': pitch,
                    'age_group': f['age_group'],
                    'priority': f['priority'],
                    'matched_pref_time': matched_time,
                    'matched_pref_pitch': matched_pitch,
                    'is_cup': f.get('is_cup', False)
                })
        
        df = pd.DataFrame(allocations) if allocations else pd.DataFrame()
        
        # ✅ ANALYZE UNALLOCATED FIXTURES
        unallocated_fixtures = set(fixtures.keys()) - allocated_fixture_ids
        
        if unallocated_fixtures:
            print(f'\n⚠️ {len(unallocated_fixtures)} fixtures could NOT be allocated:')
            
            # Categorize reasons
            reason_categories = {
                'impossible': [],  # No valid slots exist
                'capacity': [],    # Slots exist but all occupied
                'time_conflict': []  # Multiple fixtures at same time
            }
            
            for fixture_id in unallocated_fixtures:
                fdata = fixtures[fixture_id]
                fixture_date = fdata['fixture_date']
                
                # Check if fixture had any valid slots
                if fixture_id in impossible_fixtures:
                    reason_categories['impossible'].append({
                        'fixture_id': fixture_id,
                        'team': fdata['team_name'],
                        'date': fixture_date,
                        'reason': constraint_blocked.get(fixture_id, (fixture_date, ['Unknown']))[1][0] if fixture_id in constraint_blocked else 'No valid slots'
                    })
                else:
                    # Had valid slots but wasn't allocated - capacity issue
                    # Count how many slots were available
                    available_for_fixture = [v for (fid, _, _, _), v in fixture_slot_vars.items() if fid == fixture_id]
                    
                    reason_categories['capacity'].append({
                        'fixture_id': fixture_id,
                        'team': fdata['team_name'],
                        'date': fixture_date,
                        'reason': f'All {len(available_for_fixture)} compatible slots occupied'
                    })
            
            # Print categorized results
            if reason_categories['impossible']:
                print(f'\n  ❌ No compatible slots ({len(reason_categories["impossible"])} fixtures):')
                for item in reason_categories['impossible'][:10]:
                    print(f"    - {item['team']} on {item['date']}: {item['reason']}")
                if len(reason_categories['impossible']) > 10:
                    print(f"    ... and {len(reason_categories['impossible']) - 10} more")
            
            if reason_categories['capacity']:
                print(f'\n  🔒 Capacity constraints ({len(reason_categories["capacity"])} fixtures):')
                for item in reason_categories['capacity'][:10]:
                    print(f"    - {item['team']} on {item['date']}: {item['reason']}")
                if len(reason_categories['capacity']) > 10:
                    print(f"    ... and {len(reason_categories['capacity']) - 10} more")
        
        if len(df) > 0:
            # Summary statistics
            time_matches = df['matched_pref_time'].sum()
            pitch_matches = df['matched_pref_pitch'].sum()
            
            # Senior P6 allocation summary
            seniors_on_p6 = df[(df['pitch'] == 'P6 11v11 (Seniors)') & 
                               (df['age_group'].isin(['Seniors', 'Womens']))]
            
            # ✅ NEW: P3 Middle pitch allocation summary
            fixtures_on_p3 = df[df['pitch'] == 'P3 11v11 (Middle)']
            
            print(f'\n📊 Allocation Summary:')
            print(f'  - Total fixtures allocated: {len(df)}/{len(fixtures)} ({100*len(df)/len(fixtures):.1f}%)')
            print(f'  - Preferred time matches: {time_matches}/{len(df)} ({100*time_matches/len(df):.1f}%)')
            if pitch_matches > 0:
                print(f'  - Preferred pitch matches: {pitch_matches}/{len(df)} ({100*pitch_matches/len(df):.1f}%)')
            
            print(f'\n⚽ P6 Senior Pitch Allocations:')
            if len(seniors_on_p6) > 0:
                p6_by_team = seniors_on_p6.groupby('team').size().reset_index(name='fixtures')
                for _, row in p6_by_team.iterrows():
                    priority = senior_team_priority.get(row['team'], 0)
                    print(f"  - {row['team']}: {row['fixtures']} fixture(s) [priority: {priority}]")
            else:
                print(f"  - No senior fixtures on P6")
            
            # ✅ NEW: Report P3 Middle pitch usage
            if len(fixtures_on_p3) > 0:
                print(f'\n🎯 P3 11v11 (Middle) - Small Pitch Allocations:')
                p3_by_age = fixtures_on_p3.groupby('age_group').size().reset_index(name='fixtures')
                for _, row in p3_by_age.iterrows():
                    priority = p3_middle_priority.get(row['age_group'], 0)
                    priority_label = f"priority: {priority}" if priority > 0 else "no priority"
                    print(f"  - {row['age_group']}: {row['fixtures']} fixture(s) [{priority_label}]")
                
                # Show if U13/U14 got preferred allocation
                u13_u14_count = len(fixtures_on_p3[fixtures_on_p3['age_group'].isin(['U13', 'U14'])])
                if u13_u14_count > 0:
                    print(f"  ✓ {u13_u14_count}/{len(fixtures_on_p3)} allocations to priority ages (U13/U14)")
            else:
                print(f'\n🎯 P3 11v11 (Middle): No fixtures allocated')

            # ✅ NEW: Report Cup fixture allocations
            cup_fixtures = df[df['fixture_id'].isin([fid for fid, f in fixtures.items() if f.get('is_cup', False)])]
            if len(cup_fixtures) > 0:
                print(f'\n🏆 Cup Fixture Allocations:')
                print(f'  Total cup fixtures: {len(cup_fixtures)}')
                
                # Time slot analysis
                time_counts = cup_fixtures['time'].value_counts()
                for time_slot, count in sorted(time_counts.items()):
                    pct = (count / len(cup_fixtures)) * 100
                    icon = "✓" if time_slot == '09:30' else "○"
                    print(f"  {icon} {time_slot}: {count} fixtures ({pct:.0f}%)")
                
                # Show which cup fixtures got 09:30
                cup_0930 = cup_fixtures[cup_fixtures['time'] == '09:30']
                if len(cup_0930) > 0:
                    print(f'\n  Cup fixtures at preferred 09:30 slot:')
                    for _, row in cup_0930.iterrows():
                        print(f"    - {row['date']}: {row['team']} on {row['pitch']}")
                
                # Show cup fixtures that didn't get 09:30
                cup_other = cup_fixtures[cup_fixtures['time'] != '09:30']
                if len(cup_other) > 0:
                    print(f'\n  ⚠️ Cup fixtures at other times (capacity constraint):')
                    for _, row in cup_other.head(5).iterrows():
                        print(f"    - {row['date']}: {row['team']} at {row['time']} on {row['pitch']}")
                    if len(cup_other) > 5:
                        print(f"    ... and {len(cup_other) - 5} more")    
            
            # ✅ Back-to-back analysis (all days)
            print(f'\n📅 Back-to-Back Match Analysis:')
            
            total_back_to_back = 0
            all_dates = sorted(df['date'].unique())
            
            for date in all_dates:
                date_fixtures = df[df['date'] == date]
                
                # Check each pitch for back-to-back slots
                for pitch in date_fixtures['pitch'].unique():
                    pitch_fixtures = date_fixtures[date_fixtures['pitch'] == pitch]
                    times = sorted(pitch_fixtures['time'].unique())
                    
                    # Check if 09:30 AND 11:00 are both used (back-to-back)
                    if '09:30' in times and '11:00' in times:
                        total_back_to_back += 1
                        teams = pitch_fixtures.sort_values('time')['team'].tolist()
                        
                        # Check if it's a Glebelands pitch
                        pitch_info = pitches.get(pitch, {})
                        is_glebelands = pitch_info.get('location') == 'glebelands'
                        
                        if is_glebelands:
                            print(f"  ℹ️ {date} - {pitch}: Back-to-back (Glebelands overflow)")
                        else:
                            print(f"  ⚠️ {date} - {pitch}: Back-to-back matches")
                        
                        print(f"     09:30: {teams[0]}")
                        if len(teams) > 1:
                            print(f"     11:00: {teams[1]}")
            
            if total_back_to_back == 0:
                print(f"  ✓ No back-to-back matches (optimized spreading)")
            else:
                main_pitch_backtoback = len([p for p in df['pitch'].unique() 
                                            if pitches.get(p, {}).get('location') != 'glebelands'])
                print(f"  Total: {total_back_to_back} pitch(es) with back-to-back")
                print(f"  (System minimizes back-to-back but may occur if capacity is tight)")
            
            # Glebelands 3G pitch usage report
            glebelands_fixtures = df[df['pitch'].str.contains('Glebelands')]
            if len(glebelands_fixtures) > 0:
                print(f'\n🏟️ Glebelands 3G Pitch Usage:')
                for pitch in sorted(glebelands_fixtures['pitch'].unique()):
                    pitch_usage = glebelands_fixtures[glebelands_fixtures['pitch'] == pitch]
                    print(f"  {pitch}: {len(pitch_usage)} fixture(s)")
                    
                    # Show which dates used Glebelands
                    for date in sorted(pitch_usage['date'].unique()):
                        date_teams = pitch_usage[pitch_usage['date'] == date]['team'].tolist()
                        print(f"    - {date}: {', '.join(date_teams)}")
                
                # Check if Glebelands helped avoid Sunday back-to-back
                sunday_glebelands = glebelands_fixtures[glebelands_fixtures['date'].str.startswith('Sun')]
                if len(sunday_glebelands) > 0:
                    print(f"  ✓ Glebelands used on {len(sunday_glebelands['date'].unique())} Sunday(s) to avoid back-to-back matches")
            else:
                print(f'\n🏟️ Glebelands 3G: Not used (all fixtures fit on main pitches)')
            
            # Check for conflicts
            conflicts = df.groupby(['date', 'time', 'pitch']).size()
            conflicts = conflicts[conflicts > 1]
            if len(conflicts) > 0:
                print(f'\n⚠️ WARNING: {len(conflicts)} slot conflicts detected!')
            
            return df.sort_values(['date', 'time', 'pitch'])
        else:
            print('\n❌ No fixtures could be allocated')
            return None
    
    return None

# =====================================
# 📊 Visualization Functions
# =====================================
def generate_excel_schedule(df: pd.DataFrame, fixtures: Dict, output_file: str):
    """Generate an Excel workbook with formatted schedules"""
    
    # ✅ Diagnostic: Check Python environment and openpyxl
    import sys
    print(f"📊 Attempting Excel generation...")
    print(f"   Python executable: {sys.executable}")
    
    try:
        import openpyxl
        print(f"   ✓ openpyxl version: {openpyxl.__version__}")
    except ImportError as e:
        print(f"   ✗ openpyxl import failed: {e}")
        print(f"   Install in THIS environment with:")
        print(f"   {sys.executable} -m pip install openpyxl")
        return
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"   ✗ Failed to import openpyxl components: {e}")
        return
    
    print("   Creating workbook...")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Color scheme by age group
    age_colors = {
        'U7': 'FFE5E5', 'U8': 'FFE5E5',
        'U9': 'FFF3E5', 'U10': 'FFF3E5',
        'U11': 'FFFFE5', 'U12': 'FFFFE5',
        'U13': 'E5F5FF', 
        'U13G': 'E5F5FF',  # ✅ Same color as U13
        'U14': 'E5F5FF',
        'U15': 'E5FFE5', 'U16': 'E5FFE5',
        'U17': 'F5E5FF', 'U18': 'F5E5FF',
        'Seniors': 'FFD700', 'Womens': 'FFD700'
    }
    
    # ✅ Sort dates chronologically
    df_sorted = df.copy()
    df_sorted['date_sort'] = pd.to_datetime(df_sorted['date'], format='%Y-%m-%d')
    dates = df_sorted.sort_values('date_sort')['date'].unique().tolist()
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create "Summary" sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15
    
    summary_ws['A1'] = 'Cranleigh FC Pitch Allocation'
    summary_ws['A1'].font = Font(size=16, bold=True)
    summary_ws['A3'] = 'Total Fixtures Allocated:'
    summary_ws['B3'] = len(df)
    summary_ws['A4'] = 'Match Days:'
    summary_ws['B4'] = len(dates)
    summary_ws['A5'] = 'Teams:'
    summary_ws['B5'] = len(df['team'].unique())
    
    # Make summary bold
    for row in range(3, 6):
        summary_ws[f'A{row}'].font = Font(bold=True)
        summary_ws[f'B{row}'].font = Font(bold=True, size=12)
    
    # Create a sheet for each date
    for date in dates:
        # Clean sheet name (Excel has restrictions)
        sheet_name = date.replace('/', '-')[:31]
        ws = wb.create_sheet(sheet_name)
        
        date_fixtures = df[df['date'] == date]
        
        # Title
        ws['A1'] = f'Match Day: {date}'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Headers
        headers = ['Pitch', '09:30', '11:00', '14:00']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Get all pitches (show all, even if empty)
        all_pitches = list(pitches.keys())
        
        # Fill in data
        for row_idx, pitch in enumerate(all_pitches, 4):
            # Pitch name
            pitch_cell = ws.cell(row=row_idx, column=1, value=pitch)
            pitch_cell.font = Font(bold=True)
            pitch_cell.border = thin_border
            pitch_cell.alignment = Alignment(vertical='center')
            
            # Time slots
            for col_idx, time in enumerate(['09:30', '11:00', '14:00'], 2):
                fixture = date_fixtures[(date_fixtures['pitch'] == pitch) & 
                                       (date_fixtures['time'] == time)]
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                if len(fixture) > 0:
                    row_data = fixture.iloc[0]
                    team = row_data['team']
                    age = row_data['age_group']
                    
                    cell.value = f"{team}\n({age})"
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Color code by age group
                    color = age_colors.get(age, 'FFFFFF')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                else:
                    cell.value = '—'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            # Set row height
            ws.row_dimensions[row_idx].height = 30
    
    # Create "All Fixtures" sheet with filterable data
    all_ws = wb.create_sheet("All Fixtures")
    
    # Headers
    all_headers = ['Team', 'Date', 'Time', 'Pitch', 'Age Group', 'Preferred Time Match']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_idx, header in enumerate(all_headers, 1):
        cell = all_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        all_ws.cell(row=row_idx, column=1, value=row['team']).border = thin_border
        all_ws.cell(row=row_idx, column=2, value=row['date']).border = thin_border
        all_ws.cell(row=row_idx, column=3, value=row['time']).border = thin_border
        all_ws.cell(row=row_idx, column=4, value=row['pitch']).border = thin_border
        all_ws.cell(row=row_idx, column=5, value=row['age_group']).border = thin_border
        
        match_cell = all_ws.cell(row=row_idx, column=6, value='✓' if row['matched_pref_time'] else '✗')
        match_cell.border = thin_border
        match_cell.alignment = Alignment(horizontal='center')
        if row['matched_pref_time']:
            match_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            match_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Set column widths
    for col in range(1, 7):
        all_ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Add auto-filter
    all_ws.auto_filter.ref = f"A1:F{len(df) + 1}"
    
    # Create "Legend" sheet
    legend_ws = wb.create_sheet("Legend")
    legend_ws.column_dimensions['A'].width = 15
    legend_ws.column_dimensions['B'].width = 30
    
    legend_ws['A1'] = 'Age Group Colors'
    legend_ws['A1'].font = Font(size=14, bold=True)
    
    row = 3
    for age, color in sorted(age_colors.items()):
        if age in df['age_group'].values:
            legend_ws[f'A{row}'] = age
            legend_ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            legend_ws[f'A{row}'].border = thin_border
            legend_ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            count = len(df[df['age_group'] == age])
            legend_ws[f'B{row}'] = f'{count} fixtures'
            legend_ws[f'B{row}'].border = thin_border
            row += 1
    
    # Save workbook
    try:
        wb.save(output_file)
        print(f"   ✓ Excel file saved successfully")
    except Exception as e:
        print(f"   ✗ Failed to save Excel file: {e}")
        raise

def generate_html_schedule(df: pd.DataFrame, fixtures: Dict, output_file: str):
    """Generate an interactive HTML schedule visualization"""
    
    # Color scheme by age group
    age_colors = {
        'U7': '#FFE5E5', 'U8': '#FFE5E5',
        'U9': '#FFF3E5', 'U10': '#FFF3E5',
        'U11': '#FFFFE5', 'U12': '#FFFFE5',
        'U13': '#E5F5FF', 
        'U13G': '#E5F5FF',  # ✅ Same color as U13
        'U14': '#E5F5FF',
        'U15': '#E5FFE5', 'U16': '#E5FFE5',
        'U17': '#F5E5FF', 'U18': '#F5E5FF',
        'Seniors': '#FFD700', 'Womens': '#FFD700'
    }
    
    # ✅ Pitch styling - highlight Glebelands pitches differently
    pitch_styles = {
        'glebelands': 'border-left: 4px solid #10b981;'  # Green border for 3G pitches
    }
    
    # ✅ Sort dates chronologically
    # Convert date strings to datetime for proper sorting
    df_sorted = df.copy()
    df_sorted['date_sort'] = pd.to_datetime(df_sorted['date'], format='%Y-%m-%d')
    dates = df_sorted.sort_values('date_sort')['date'].unique().tolist()
    
    html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cranleigh FC Pitch Allocation Schedule</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            background: #f5f5f5;
            padding: 20px;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            padding: 30px;
        }
        h1 {
            color: #1a1a1a;
            margin-bottom: 10px;
            font-size: 28px;
        }
        .subtitle {
            color: #666;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .date-tabs {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            overflow-x: auto;
            padding-bottom: 10px;
        }
        .date-tab {
            padding: 10px 20px;
            background: #f0f0f0;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            white-space: nowrap;
            font-size: 14px;
            transition: all 0.2s;
        }
        .date-tab:hover {
            background: #e0e0e0;
        }
        .date-tab.active {
            background: #2563eb;
            color: white;
        }
        .schedule-grid {
            display: none;
            overflow-x: auto;
        }
        .schedule-grid.active {
            display: block;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border: 1px solid #e0e0e0;
        }
        th {
            background: #f8f9fa;
            font-weight: 600;
            color: #1a1a1a;
            position: sticky;
            top: 0;
        }
        .pitch-name {
            font-weight: 600;
            color: #1a1a1a;
        }
        .glebelands-pitch {
            background: #f0fdf4 !important;
            border-left: 4px solid #10b981;
        }
        .glebelands-label {
            display: inline-block;
            background: #10b981;
            color: white;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 10px;
            margin-left: 5px;
            font-weight: 600;
        }
        .fixture {
            padding: 8px;
            border-radius: 4px;
            font-size: 13px;
            line-height: 1.4;
        }
        .team-name {
            font-weight: 600;
            margin-bottom: 2px;
        }
        .age-badge {
            display: inline-block;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 11px;
            font-weight: 600;
            background: rgba(0,0,0,0.1);
        }
        .empty-slot {
            color: #999;
            font-style: italic;
            text-align: center;
        }
        .legend {
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
            margin-top: 20px;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 6px;
        }
        .legend-item {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 12px;
        }
        .legend-color {
            width: 20px;
            height: 20px;
            border-radius: 3px;
            border: 1px solid #ddd;
        }
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            padding: 15px;
            background: #f8f9fa;
            border-radius: 6px;
            border-left: 4px solid #2563eb;
        }
        .stat-value {
            font-size: 24px;
            font-weight: 700;
            color: #1a1a1a;
            margin-bottom: 5px;
        }
        .stat-label {
            font-size: 12px;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        @media print {
            .date-tabs { display: none; }
            .schedule-grid { display: block !important; page-break-after: always; }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>⚽ Cranleigh FC Pitch Allocation Schedule</h1>
        <p class="subtitle">Season fixtures and pitch assignments</p>
        
        <div class="stats">
            <div class="stat-card">
                <div class="stat-value">""" + str(len(df)) + """</div>
                <div class="stat-label">Total Fixtures Allocated</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">""" + str(len(dates)) + """</div>
                <div class="stat-label">Match Days</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">""" + str(len(df['team'].unique())) + """</div>
                <div class="stat-label">Teams</div>
            </div>
        </div>
        
        <div class="date-tabs">
"""
    
    # Generate date tabs
    for i, date in enumerate(dates):
        active = "active" if i == 0 else ""
        html += f'            <button class="date-tab {active}" onclick="showDate(\'{date}\')">{date}</button>\n'
    
    html += """        </div>
"""
    
    # Generate schedule for each date
    for i, date in enumerate(dates):
        active = "active" if i == 0 else ""
        date_fixtures = df[df['date'] == date]
        
        html += f"""
        <div class="schedule-grid {active}" id="date-{date}">
            <h2 style="margin-bottom: 15px; color: #1a1a1a;">{date}</h2>
            <table>
                <thead>
                    <tr>
                        <th style="width: 180px;">Pitch</th>
                        <th>09:30</th>
                        <th>11:00</th>
                        <th>14:00</th>
                    </tr>
                </thead>
                <tbody>
"""
        
        # Get all pitches that have fixtures this date
        pitches_used = sorted(date_fixtures['pitch'].unique())
        if not pitches_used:
            pitches_used = list(pitches.keys())[:5]  # Show some pitches even if empty
        
        for pitch in pitches_used:
            html += f'                    <tr>\n'
            
            # Check if this is a Glebelands pitch
            pitch_info = pitches.get(pitch, {})
            is_glebelands = pitch_info.get('location') == 'glebelands'
            pitch_class = ' class="glebelands-pitch"' if is_glebelands else ''
            
            pitch_display = pitch
            if is_glebelands:
                pitch_display = f'{pitch} <span class="glebelands-label">3G</span>'
            
            html += f'                        <td{pitch_class}><span class="pitch-name">{pitch_display}</span></td>\n'
            
            for time in ['09:30', '11:00', '14:00']:
                fixture = date_fixtures[(date_fixtures['pitch'] == pitch) & (date_fixtures['time'] == time)]
                
                if len(fixture) > 0:
                    row = fixture.iloc[0]
                    age = row['age_group']
                    color = age_colors.get(age, '#f0f0f0')
                    
                    html += f'''                        <td style="background: {color};">
                            <div class="fixture">
                                <div class="team-name">{row['team']}</div>
                                <span class="age-badge">{age}</span>
                            </div>
                        </td>
'''
                else:
                    html += '                        <td><div class="empty-slot">—</div></td>\n'
            
            html += '                    </tr>\n'
        
        html += """                </tbody>
            </table>
        </div>
"""
    
    # Add legend
    html += """
        <div class="legend">
            <strong style="width: 100%; margin-bottom: 5px;">Age Groups:</strong>
"""
    
    for age, color in sorted(age_colors.items()):
        if age in df['age_group'].values:
            html += f'            <div class="legend-item"><div class="legend-color" style="background: {color};"></div><span>{age}</span></div>\n'
    
    html += """        </div>
    </div>
    
    <script>
        function showDate(date) {
            // Hide all schedule grids
            document.querySelectorAll('.schedule-grid').forEach(grid => {
                grid.classList.remove('active');
            });
            
            // Show selected date
            document.getElementById('date-' + date).classList.add('active');
            
            // Update active tab
            document.querySelectorAll('.date-tab').forEach(tab => {
                tab.classList.remove('active');
            });
            event.target.classList.add('active');
        }
    </script>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

# =====================================
# 🚀 Main Execution
# =====================================
if __name__ == '__main__':
    try:
        fixtures, slots_by_date = load_and_validate_fixtures('cranleigh_home_fixtures.csv')
        result = solve_allocation(fixtures, slots_by_date, timeout=30)
        
        if result is not None:
            # Export CSV
            output_file = 'pitch_allocations_fixed.csv'
            result.to_csv(output_file, index=False)
            print(f'\n✅ Exported to {output_file}')
            
            # ✅ Generate HTML visualization
            html_file = 'pitch_allocations_schedule.html'
            generate_html_schedule(result, fixtures, html_file)
            print(f'✅ Generated interactive schedule: {html_file}')
            
            # ✅ Generate Excel workbook
            excel_file = 'pitch_allocations_schedule.xlsx'
            try:
                generate_excel_schedule(result, fixtures, excel_file)
                print(f'✅ Generated Excel workbook: {excel_file}')
            except Exception as e:
                print(f'⚠️ Excel generation failed: {e}')
                print(f'   Make sure openpyxl is installed: pip install openpyxl')
                import traceback
                traceback.print_exc()
            
            # ✅ Auto-open HTML in browser
            try:
                html_path = os.path.abspath(html_file)
                webbrowser.open(f'file://{html_path}')
                print(f'🌐 Opening schedule in browser...')
            except Exception as e:
                print(f'⚠️ Could not auto-open browser: {e}')
                print(f'   Please manually open: {html_file}')
            
            # Display sample
            print('\n📋 Sample allocations:')
            print(result[['team', 'date', 'time', 'pitch', 'age_group']].head(10))
            
            # Show allocation by date with comparison
            print('\n📅 Fixtures allocated vs expected per date:')
            actual_by_date = result.groupby('date').size()
            
            # Get expected counts from original CSV
            fixtures_df = pd.read_csv('cranleigh_home_fixtures.csv')
            expected_by_date = fixtures_df.groupby('match_date').size()
            
            comparison = pd.DataFrame({
                'Expected': expected_by_date,
                'Allocated': actual_by_date
            }).fillna(0).astype(int)
            comparison['Missing'] = comparison['Expected'] - comparison['Allocated']
            
            for date, row in comparison.iterrows():
                status = "✅" if row['Missing'] == 0 else "⚠️"
                print(f"  {status} {date}: {row['Allocated']}/{row['Expected']} allocated", end="")
                if row['Missing'] > 0:
                    print(f" ({row['Missing']} missing)")
                else:
                    print()
            
            # Show which fixtures are missing
            allocated_fixture_ids = set(result['fixture_id'])
            all_fixture_ids = set(fixtures.keys())
            missing_fixtures = all_fixture_ids - allocated_fixture_ids
            
            if missing_fixtures:
                print(f"\n⚠️ {len(missing_fixtures)} fixtures NOT allocated:")
                for fid in sorted(list(missing_fixtures))[:10]:
                    print(f"  - {fid}")
                if len(missing_fixtures) > 10:
                    print(f"  ... and {len(missing_fixtures) - 10} more")
        else:
            print('\n❌ No solution found')
            print("\nPossible reasons:")
            print("  1. Too many fixtures for available slots")
            print("  2. Conflicting time constraints")
            print("  3. Insufficient pitches of required formats")
            
    except Exception as e:
        print(f'\n❌ Error: {e}')
        import traceback
        traceback.print_exc()

